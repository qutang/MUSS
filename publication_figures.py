import numpy as np
import pandas as pd
import os
import matplotlib.pyplot as plt
from matplotlib import rcParams
import matplotlib.gridspec as gridspec
import seaborn as sns
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, PatternFill
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter


def format_for_excel(df, highlight_header=True):
    df = df.round(3)
    wb = Workbook()
    ws = wb.active
    font_style = Font(name='Times New Roman', size=12)
    redFill = PatternFill(start_color='EE1111',
                          end_color='EE1111',
                          fill_type='solid')
    greenFill = PatternFill(start_color='11EE11',
                            end_color='11EE11',
                            fill_type='solid')
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)

    # whole sheet style
    for row in ws.iter_rows():
        for cell in row:
            cell.font = font_style
            # first row
            if cell.row == 1:
                cell.font = cell.font + Font(bold=True)
                cell.border = Border(
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                if cell.column == 'A':
                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
            # first column
            elif cell.column == 'A':
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin')
                )

    for col in range(2, ws.max_column + 1):
        col_letter = get_column_letter(col)
        ws.conditional_formatting.add(col_letter + '2:' + col_letter + str(ws.max_row), FormulaRule(
            formula=[col_letter + '2>=LARGE($' + col_letter + '$2:$' + col_letter + '$' + str(ws.max_row) + ',5)'], fill=redFill))
        ws.conditional_formatting.add(col_letter + '2:' + col_letter + str(ws.max_row), FormulaRule(
            formula=[col_letter + '2<=SMALL($' + col_letter + '$2:$' + col_letter + '$' + str(ws.max_row) + ',5)'], fill=greenFill))

    return wb


def table_3(summary_file):
    output_filepath = summary_file.replace(
        'prediction_sets', 'publication_figures_and_tables').replace('summary.csv', 'table3.csv')
    output_filepath_excel = summary_file.replace(
        'prediction_sets', 'publication_figures_and_tables').replace('summary.csv', 'table3.xlsx')
    os.makedirs(os.path.dirname(output_filepath), exist_ok=True)
    summary = pd.read_csv(summary_file)
    filter_condition = (summary['FEATURE_TYPE'] == 'MO') & (
        summary['NUM_OF_SENSORS'] == 2)
    table3_data = summary.loc[filter_condition, [
        'SENSOR_PLACEMENT', 'POSTURE_AVERAGE', 'LYING_POSTURE', 'SITTING_POSTURE', 'UPRIGHT_POSTURE']]
    table3_data.columns = ['Sensor placements',
                           'Average', 'Lying', 'Sitting', 'Upright']
    table3_data = table3_data.sort_values(by=['Average'], ascending=False)
    table3_data.loc[:, 'Sensor placements'] = table3_data['Sensor placements'].transform(
        lambda s: s.replace('_', ', '))
    table3_wb = format_for_excel(table3_data)
    table3_data.to_csv(output_filepath, float_format='%.2f', index=False)
    table3_wb.save(output_filepath_excel)


def table_4(summary_file):
    output_filepath = summary_file.replace(
        'prediction_sets', 'publication_figures_and_tables').replace('summary.csv', 'table4.csv')
    output_filepath_excel = summary_file.replace(
        'prediction_sets', 'publication_figures_and_tables').replace('summary.csv', 'table4.xlsx')
    os.makedirs(os.path.dirname(output_filepath), exist_ok=True)
    summary = pd.read_csv(summary_file)
    filter_condition = (summary['FEATURE_TYPE'] == 'MO') & (
        summary['NUM_OF_SENSORS'] == 2)
    table4_data = summary.loc[filter_condition, [
        'SENSOR_PLACEMENT', 'ACTIVITY_AVERAGE',  'ACTIVITY_GROUP_AVERAGE', 'ACTIVITY_IN_GROUP_AVERAGE']]
    table4_data.columns = ['Sensor placements',
                           'Average', 'Between activity groups', 'Within activity groups']
    table4_data = table4_data.sort_values(by=['Average'], ascending=False)
    table4_data.loc[:, 'Sensor placements'] = table4_data['Sensor placements'].transform(
        lambda s: s.replace('_', ', '))
    table4_wb = format_for_excel(table4_data)
    table4_data.to_csv(output_filepath, float_format='%.2f', index=False)
    table4_wb.save(output_filepath_excel)


def figure_1(summary_file):
    rcParams['font.family'] = 'serif'
    rcParams['font.size'] = 12
    rcParams['font.serif'] = ['Times New Roman']
    # setup configurations
    figure_file_extensions = ['.png', '.svg', '.pdf', '.eps']
    output_filepaths = [summary_file.replace(
        'prediction_sets', 'publication_figures_and_tables').replace('summary.csv', 'figure1' + extension) for extension in figure_file_extensions]
    os.makedirs(os.path.dirname(output_filepaths[0]), exist_ok=True)

    # read data
    summary = pd.read_csv(summary_file)

    #  prepare line plot data
    line_plot_data = summary[[
        'NUM_OF_SENSORS', 'FEATURE_TYPE', 'POSTURE_AVERAGE', 'ACTIVITY_AVERAGE']]
    line_plot_data.columns = [
        'Number of sensors', 'Feature set', 'Posture', 'PA']
    line_plot_data.loc[:, 'Feature set'] = line_plot_data['Feature set'].map({
        'M': 'Motion features only',
        'O': 'Orientation related features only',
        'MO': 'Motion + orientation related features'
    })

    # prepare point plot data
    point_plot_data = summary.loc[summary['FEATURE_TYPE'] == 'MO', [
        'SENSOR_PLACEMENT', 'NUM_OF_SENSORS', 'POSTURE_AVERAGE', 'ACTIVITY_AVERAGE']]
    point_plot_data.columns = ['Sensor placements',
                               'Number of sensors', 'Posture', 'PA']

    point_plot_data = pd.melt(point_plot_data, id_vars=['Sensor placements', 'Number of sensors'], value_vars=[
                              'Posture', 'PA'], var_name='Classification task', value_name='F1-score')
    point_plot_data['Include dominant wrist'] = point_plot_data['Sensor placements'].transform(
        lambda s: 'DW' in s.split('_'))
    point_plot_data['Include nondominant wrist'] = point_plot_data['Sensor placements'].transform(
        lambda s: 'NDW' in s.split('_'))
    point_plot_data['Include wrists'] = point_plot_data['Include dominant wrist'] | point_plot_data['Include nondominant wrist']

    point_plot_data = point_plot_data.sort_values(
        by=['Number of sensors', 'Classification task', 'F1-score'], ascending=True)

    point_plot_data = point_plot_data.groupby(['Number of sensors', 'Classification task']).apply(
        lambda rows: pd.concat((rows.head(5), rows.tail(5))))
    point_plot_data = point_plot_data.reset_index(drop=True).drop_duplicates()

    # draw plots
    g, axes = plt.subplots(2, 2, figsize=(8, 8))
    sns.set_context("paper")
    for task, index in zip(['Posture', 'PA'], [0, 1]):
        sns.set(rc={"lines.linewidth": 0.8,
                "font.family": ['serif'],
                "font.serif": ['Times New Roman'],
                "font.size": 12
                })
        # draw swarm and line for MO feature set
        axes[index][0].set_ylim(0, 1.2)
        axes[index][0].set_xlim(0, 7)
        axes[index][0].yaxis.grid(linestyle='--')
        swarm_data = point_plot_data.loc[point_plot_data['Classification task'] == task, :]
        if task == 'Posture':
            swarm_data_wrist = swarm_data.loc[swarm_data['Include wrists'] == True, :]
            swarm_data_nonwrist = swarm_data.loc[swarm_data['Include wrists'] == False, :]
        else:
            swarm_data_wrist = swarm_data.loc[swarm_data['Include dominant wrist'] == True, :]
            swarm_data_nonwrist = swarm_data.loc[swarm_data['Include dominant wrist'] == False, :]
        line_data_mo = line_plot_data.loc[line_plot_data['Feature set'] == 'Motion + orientation related features', [
            'Number of sensors', 'Feature set', task]].rename(columns={task: 'F1-score'})
        for x in [0.5, 1.5, 2.5, 3.5, 4.5, 5.5, 6.5]:
            axes[index][0].axvline(x=x, color='0.75')
        sns.swarmplot(x = 'Number of sensors', y = 'F1-score', data=swarm_data, ax=axes[index][0], linewidth=1, hue='Include wrists', palette='Greys', size=4)
        sns.pointplot(x='Number of sensors', y='F1-score',
                      data=line_data_mo, ax=axes[index][0], color='gray', marker='x', capsize=0.1, errwidth=0, hue='Feature set')
        legend_handles = axes[index][0].legend_.legendHandles
        legend_handles[2] = axes[index][0].lines[7]
        if task == 'Posture':
            axes[index][0].legend(handles=legend_handles, labels=["Models include W sensors", "Models not include W sensors", "M + O features"], frameon=True, loc='lower right', framealpha=1, fancybox=False, facecolor='white', edgecolor='black', shadow=None)
        else:
            axes[index][0].legend(handles=legend_handles, labels=["Models include DW sensors", "Models not include DW sensors", "M + O features"], frameon=True, loc='lower right', framealpha=1, fancybox=False, facecolor='white', edgecolor='black', shadow=None)

        # draw line for other feature set

        line_data_others = line_plot_data.loc[line_plot_data['Feature set'] != 'Motion + orientation related features', [
            'Number of sensors', 'Feature set', task]].rename(columns={task: 'F1-score'})
        sns.pointplot(x='Number of sensors', y='F1-score', data=line_data_others,
                      dodge=True, ax=axes[index][1], hue='Feature set', palette='Greys', linestyles=['--', '-.'], markers='x', errwidth=0)
        axes[index][1].legend(handles=[axes[index][1].lines[0], axes[index][1].lines[8]], labels=["M features","O features"], frameon=True, loc='lower right', framealpha=1, fancybox=False, facecolor='white', edgecolor='black', shadow=None)
        axes[index][1].set_ylim(0, 1.2)
        axes[index][1].set_yticklabels([])
        axes[index][1].yaxis.set_major_formatter(plt.NullFormatter())
        axes[index][1].set_ylabel('')
        axes[index][1].yaxis.grid(linestyle='--')
        axes[index][1].spines['left'].set_color('grey')

    g.subplots_adjust(wspace=0, hspace=0.35)
    plt.figtext(0.5, 0.49, '(a) Posture recognition performances',
                ha='center', va='top')
    plt.figtext(0.5, 0.05, '(b) PA recognition performances',
                ha='center', va='top')
    # plt.show()
    # save figure in different formats
    for output_filepath in output_filepaths:
        plt.savefig(output_filepath, dpi=300, orientation='landscape')


if __name__ == '__main__':
    dataset_folder = 'D:/data/spades_lab/'
    summary_file = os.path.join(
        dataset_folder, 'DerivedCrossParticipants', 'location_matters', 'prediction_sets', 'summary.csv')
    # table_3(summary_file)
    # table_4(summary_file)
    figure_1(summary_file)
